import xlrd
import xlwt
import os
import openpyxl
import numpy as np
import regex

filelist1 = []
filelist2 = []
Year = []
Value1 = []
Value2 = []

def follow_up():
    col = 1
    row = 1
    max = 1
    while row < 50 and col < 50:
        a = sheet.cell(row, col).value
        if a>max:
            return a
        else:
            return max
        row+=1
        col+=1

# print(follow_up())

for i in range(2001,2018): 
    pat = "/Users/ps/Documents/Transsight/BART-Open Data Portal/Datasets/Datasets for Upload/Ridership/ridership_" + str(i)
    for filename in os.listdir(pat):

      if filename.endswith('Entry_Exit Matrices.xls'):
        # print('yes' + filename)
        filelist2.append(filename[:-24])
        pathstr = "/Users/ps/Documents/Transsight/BART-Open Data Portal/Datasets/Datasets for Upload/Ridership/ridership_" + str(i)+ "/" + filename 
        wb = xlrd.open_workbook(pathstr)
        sheet = wb.sheet_by_index(0)
        print(filename)
        c = sheet.ncols
        r = sheet.nrows
        print(c)
        print(r)
        cnt = 0
        for x in range(0, c+1):
          try: 
            if sheet.cell_value(1, x) == "Exits":
              print("yes")
              cnt = cnt + 1
              print ("column: ", x)
              print("row: ", r)
              print(cnt)          
              ips = sheet.cell_value(rowx = r-1, colx = x)
              print(ips)
            break
          except IndexError:
#             continue
            Value2.append(ips)
            print (Value2)

      elif filename.startswith('Ridership_') and filename.endswith('.xls'): 
    # or filename.endswith('.xls'):
        filelist1.append(filename[10:-4])
        pathstr = "/Users/ps/Documents/Transsight/BART-Open Data Portal/Datasets/Datasets for Upload/Ridership/ridership_" + str(i)+ "/" + filename 
        wb = xlrd.open_workbook(pathstr)
        sheet = wb.sheet_by_index(0)
        Value2.append(sheet.cell_value(rowx = 45, colx = 44))

      elif filename.startswith('Ridership_') and filename.endswith('.xlsx'): 
        filelist1.append(filename[10:-5])
        pathstr = "/Users/ps/Documents/Transsight/BART-Open Data Portal/Datasets/Datasets for Upload/Ridership/ridership_" + str(i)+ "/" + filename 
        wb = xlrd.open_workbook(pathstr)
        sheet = wb.sheet_by_index(0)
        Value2.append(sheet.cell_value(rowx = 45, colx = 44))  
     

      
        

    
        # print(filelist1)
      
        # print(filelist)
        # Year.append(filename[10:-5])
        # pathstr = "/Users/ps/Documents/Transsight/BART-Open Data Portal/Datasets/Datasets for Upload/Ridership/ridership_" + str(i)+ "/" + filename 
        # wb = xlrd.open_workbook(pathstr)
        # sheet = wb.sheet_
        # Value.append(sheet.cell_value(rowx = 41, colx = 41))
        # print (Value)
filelist  = filelist2+filelist1
Value = Value1 + Value2

# print("FILELIST..............")
# print (filelist)
# print (Value)

# # print("FILELIST2..............")
# # print (filelist2)
# # print (Year)
# # print(Value)
# # print(len(filelist))

import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
import numpy as np
 
df = pd.DataFrame({'Month':filelist,
                   'Ridership': Value})
writer = pd.ExcelWriter('/Users/ps/Documents/totall.xlsx')
df.to_excel(writer,'Sheet1',index=False)
writer.save()




# p = "Users/ps/Documents/Transsight/BART-Open Data Portal/Datasets/Datasets for Upload/Ridership/ridership_2017/Ridership_April2017.xlsx"
# wb = openpyxl.load_workbook("/Users/ps/Documents/Transsight/BART-Open Data Portal/Datasets/Datasets for Upload/Ridership/ridership_2017/Ridership_April2017.xlsx")
# sheet = wb.get_sheet_by_name("Weekday OD")
# # print(sheet['A3'].value)
# v = sheet.cell(row = 49, column = 48).value
# print(os.path.basename(p)[10:-5])
# print(v)
# book = xlwt.Workbook(encoding = "utf-8")
# sheet1 = book.add_sheet("Ridership")

#----------------------------------------
# get data from excel file
#----------------------------------------
# XLS_FILE = "C:\\desktop\\test.xls"
# ROW_SPAN = (1, 16)
# COL_SPAN = (1, 6)
# app = Dispatch("Excel.Application")
# app.Visible = True
# ws = app.Workbooks.Open(XLS_FILE).Sheets(1)
# xldata = [[ws.Cells(row, col).Value 
#               for col in xrange(COL_SPAN[0], COL_SPAN[1])] 
#              for row in xrange(ROW_SPAN[0], ROW_SPAN[1])]
# #print xldata
# a = np.asarray(list(xldata), dtype='object')
# print(a)

