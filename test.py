
import xlrd
import xlwt
import os
import openpyxl
import numpy as np
filelist = []
Year = []
Value = []

for i in range(2001,2018): 
    pat = "/Users/ps/Documents/Transsight/BART-Open Data Portal/Datasets/Datasets for Upload/Ridership/ridership_" + str(i)
    for filename in os.listdir(pat):
      if filename.startswith('Ridership_') and filename.endswith('.xlsx'): 
    # or filename.endswith('.xls'):
        print('no')
        filelist.append(filename[:-24])
        print(filelist)
        
      elif filename.endswith('Entry_Exit Matrices.xls'):
        print('yes' + filename)
        filelist.append(filename[10:-5])
        # print(filelist)
        Year.append(filename[-9:-5])
        pathstr = "/Users/ps/Documents/Transsight/BART-Open Data Portal/Datasets/Datasets for Upload/Ridership/ridership_" + str(i)+ "/" + filename 
        wb = xlrd.open_workbook(pathstr)
        sheet = wb.sheet_by_index(0)
        # Value.append(sheet.cell(row = 49, column = 48).value)
        Value.append(sheet.cell_value(rowx = 0, colx = 0))

# print (filelist)
# print (Year)
# print(Value)



# print(len(filelist))

import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
import numpy as np
 
df = pd.DataFrame({'Month':filelist,
                   'Year':Year,
                   'Ridership': Value})
writer = pd.ExcelWriter('/Users/ps/Documents/totall.xlsx')
df.to_excel(writer,'Sheet1',index=False)
writer.save()




# p = "Users/ps/Documents/Transsight/BART-Open Data Portal/Datasets/Datasets for Upload/Ridership/ridership_2017/Ridership_April2017.xlsx"
# wb = openpyxl.load_workbook("/Users/ps/Documents/Transsight/BART-Open Data Portal/Datasets/Datasets for Upload/Ridership/ridership_2017/Ridership_April2017.xlsx")
# sheet = wb.get_sheet_by_name("Weekday OD")
# # print(sheet['A3'].value)
# v = sheet.cell(row = 49, column = 48).value
# print(os.path.basename(p)[10:-5])
# print(v)
# book = xlwt.Workbook(encoding = "utf-8")
# sheet1 = book.add_sheet("Ridership")

#----------------------------------------
# get data from excel file
#----------------------------------------
# XLS_FILE = "C:\\desktop\\test.xls"
# ROW_SPAN = (1, 16)
# COL_SPAN = (1, 6)
# app = Dispatch("Excel.Application")
# app.Visible = True
# ws = app.Workbooks.Open(XLS_FILE).Sheets(1)
# xldata = [[ws.Cells(row, col).Value 
#               for col in xrange(COL_SPAN[0], COL_SPAN[1])] 
#              for row in xrange(ROW_SPAN[0], ROW_SPAN[1])]
# #print xldata
# a = np.asarray(list(xldata), dtype='object')
# print(a)

